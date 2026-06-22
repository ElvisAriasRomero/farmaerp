"""Generación de archivos de reporte (PDF / Excel) a partir de datos reales.

Cada tipo de reporte arma una estructura genérica:
    {
        "titulo": str,
        "subtitulo": str,
        "headers": [str, ...],
        "rows": [[val, ...], ...],
        "aligns": ["left"|"right", ...],   # opcional
        "total": ("Etiqueta", "valor")    # opcional, fila resaltada al pie
    }
y luego se renderiza a PDF (reportlab) o Excel (openpyxl).
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum


# --------------------------------------------------------------------------- #
#  Helpers de formato
# --------------------------------------------------------------------------- #
def _money(v):
    if v is None:
        return "—"
    return f"Bs {Decimal(v):,.2f}"


def _fdate(d):
    return d.strftime("%d/%m/%Y") if d else "—"


def _rango_texto(reporte):
    if reporte.fecha_inicio and reporte.fecha_fin:
        return f"Período: {_fdate(reporte.fecha_inicio)} – {_fdate(reporte.fecha_fin)}"
    if reporte.fecha_inicio:
        return f"Desde: {_fdate(reporte.fecha_inicio)}"
    if reporte.fecha_fin:
        return f"Hasta: {_fdate(reporte.fecha_fin)}"
    return "Período: completo"


# --------------------------------------------------------------------------- #
#  Constructores de datos por tipo de reporte
# --------------------------------------------------------------------------- #
def _data_ventas(reporte):
    from apps.ventas.models import Venta

    qs = Venta.objects.select_related("cliente", "empleado").all()
    if reporte.fecha_inicio:
        qs = qs.filter(fecha_venta__date__gte=reporte.fecha_inicio)
    if reporte.fecha_fin:
        qs = qs.filter(fecha_venta__date__lte=reporte.fecha_fin)
    qs = qs.order_by("id_venta")

    rows = []
    total = Decimal("0")
    for v in qs:
        rows.append([
            f"#{v.id_venta}",
            _fdate(v.fecha_venta),
            v.cliente.nombre if v.cliente else "Consumidor final",
            v.empleado.nombre if v.empleado else "—",
            "Sí" if v.con_factura else "No",
            v.estado.capitalize(),
            _money(v.total),
        ])
        total += v.total or 0
    return {
        "titulo": "Reporte de Ventas",
        "headers": ["N°", "Fecha", "Cliente", "Vendedor", "Factura", "Estado", "Total"],
        "aligns": ["left", "left", "left", "left", "left", "left", "right"],
        "rows": rows,
        "total": ("Total vendido", _money(total)),
    }


def _data_compra(reporte):
    from apps.compras.models import Compra

    qs = Compra.objects.select_related("proveedor", "empleado").all()
    if reporte.fecha_inicio:
        qs = qs.filter(fecha_pedido__gte=reporte.fecha_inicio)
    if reporte.fecha_fin:
        qs = qs.filter(fecha_pedido__lte=reporte.fecha_fin)
    qs = qs.order_by("id_compra")

    rows = []
    total = Decimal("0")
    for c in qs:
        rows.append([
            f"#{c.id_compra}",
            _fdate(c.fecha_pedido),
            _fdate(c.fecha_recepcion),
            c.proveedor.nombre if c.proveedor else "—",
            c.estado.capitalize(),
            _money(c.total),
        ])
        total += c.total or 0
    return {
        "titulo": "Reporte de Compras",
        "headers": ["N°", "Pedido", "Recepción", "Proveedor", "Estado", "Total"],
        "aligns": ["left", "left", "left", "left", "left", "right"],
        "rows": rows,
        "total": ("Total comprado", _money(total)),
    }


def _data_inventario(reporte):
    from apps.productos.models import Producto

    qs = Producto.objects.select_related("categoria", "inventario").order_by("nombre")
    rows = []
    valor_total = Decimal("0")
    for p in qs:
        stock = getattr(getattr(p, "inventario", None), "stock_actual", 0) or 0
        minimo = getattr(getattr(p, "inventario", None), "stock_minimo", 0) or 0
        estado = "Sin stock" if stock <= 0 else ("Bajo" if stock < minimo else "OK")
        rows.append([
            p.nombre,
            p.categoria.nombre if p.categoria else "—",
            str(stock),
            str(minimo),
            estado,
            _money(p.precio_compra),
            _money(p.precio_venta),
        ])
        if p.precio_compra:
            valor_total += Decimal(p.precio_compra) * stock
    return {
        "titulo": "Reporte de Inventario",
        "headers": ["Producto", "Categoría", "Stock", "Mínimo", "Estado", "Costo", "Venta"],
        "aligns": ["left", "left", "right", "right", "left", "right", "right"],
        "rows": rows,
        "total": ("Valor del inventario (a costo)", _money(valor_total)),
    }


def _data_auditoria(reporte):
    from apps.auditoria.models import BitacoraAuditoria

    qs = BitacoraAuditoria.objects.select_related("empleado").all()
    if reporte.fecha_inicio:
        qs = qs.filter(fecha_operacion__gte=reporte.fecha_inicio)
    if reporte.fecha_fin:
        qs = qs.filter(fecha_operacion__lte=reporte.fecha_fin)
    qs = qs.order_by("id_bitacora")

    rows = []
    for b in qs:
        rows.append([
            _fdate(b.fecha_operacion),
            b.hora_operacion.strftime("%H:%M") if b.hora_operacion else "—",
            b.empleado.nombre if b.empleado else "—",
            b.tabla_afectada,
            b.tipo_operacion.capitalize(),
        ])
    return {
        "titulo": "Reporte de Auditoría",
        "headers": ["Fecha", "Hora", "Empleado", "Tabla", "Operación"],
        "aligns": ["left", "left", "left", "left", "left"],
        "rows": rows,
    }


def _data_prediccion(reporte):
    rows = []
    try:
        from apps.prediccion.models import Prediccion

        qs = Prediccion.objects.select_related("producto").order_by("-id_prediccion")[:100]
        for p in qs:
            rows.append([
                getattr(getattr(p, "producto", None), "nombre", "—"),
                str(getattr(p, "cantidad_estimada", getattr(p, "demanda_estimada", "—"))),
                _fdate(getattr(p, "fecha_objetivo", None)),
            ])
    except Exception:
        rows = []
    return {
        "titulo": "Reporte de Predicción de Demanda",
        "headers": ["Producto", "Demanda estimada", "Fecha objetivo"],
        "aligns": ["left", "right", "left"],
        "rows": rows,
    }


_BUILDERS = {
    "ventas": _data_ventas,
    "compra": _data_compra,
    "inventario": _data_inventario,
    "auditoria": _data_auditoria,
    "prediccion": _data_prediccion,
}


def construir_datos(reporte):
    builder = _BUILDERS.get(reporte.tipo)
    if not builder:
        return {
            "titulo": f"Reporte {reporte.tipo}",
            "headers": ["Sin datos"],
            "rows": [],
        }
    return builder(reporte)


# --------------------------------------------------------------------------- #
#  Renderizadores
# --------------------------------------------------------------------------- #
def _render_pdf(data, reporte):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = BytesIO()
    many_cols = len(data["headers"]) > 5
    pagesize = landscape(A4) if many_cols else A4
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=16 * mm, rightMargin=16 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
        title=data["titulo"],
    )
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontSize=17,
                       textColor=colors.HexColor("#1e3a8a"), spaceAfter=2)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9.5,
                         textColor=colors.HexColor("#64748b"))

    elems = [Paragraph("FarmaERP — Gestión farmacéutica", sub),
             Paragraph(data["titulo"], h)]
    meta = _rango_texto(reporte)
    gen = reporte.empleado.nombre if reporte.empleado_id and reporte.empleado else "—"
    elems.append(Paragraph(
        f"{meta} &nbsp;·&nbsp; Generado por: {gen} &nbsp;·&nbsp; "
        f"Emitido: {date.today().strftime('%d/%m/%Y')}", sub))
    elems.append(Spacer(1, 8 * mm))

    table_rows = [data["headers"]] + (data["rows"] or [["Sin registros en el período."] +
                                      [""] * (len(data["headers"]) - 1)])
    tbl = Table(table_rows, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#0f172a")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor("#1e3a8a")),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    for i, al in enumerate(data.get("aligns", [])):
        style.append(("ALIGN", (i, 0), (i, -1), "RIGHT" if al == "right" else "LEFT"))
    tbl.setStyle(TableStyle(style))
    elems.append(tbl)

    if data.get("total"):
        elems.append(Spacer(1, 6 * mm))
        tot = Table([[data["total"][0], data["total"][1]]],
                    colWidths=[doc.width - 45 * mm, 45 * mm])
        tot.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eff6ff")),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10.5),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#1e3a8a")),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        elems.append(tot)

    doc.build(elems)
    buf.seek(0)
    return buf.getvalue()


def _render_excel(data, reporte):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = data["titulo"][:31]

    head_fill = PatternFill("solid", fgColor="1E3A8A")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=14, color="1E3A8A")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols = len(data["headers"])

    ws.cell(row=1, column=1, value=data["titulo"]).font = title_font
    gen = reporte.empleado.nombre if reporte.empleado_id and reporte.empleado else "—"
    ws.cell(row=2, column=1,
            value=f"{_rango_texto(reporte)} · Generado por: {gen} · "
                  f"Emitido: {date.today().strftime('%d/%m/%Y')}").font = Font(
        size=9, color="64748B")

    hrow = 4
    for j, hcell in enumerate(data["headers"], start=1):
        c = ws.cell(row=hrow, column=j, value=hcell)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    r = hrow + 1
    for row in data["rows"]:
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = border
            al = data.get("aligns", [])
            if j - 1 < len(al) and al[j - 1] == "right":
                c.alignment = Alignment(horizontal="right")
        r += 1

    if not data["rows"]:
        ws.cell(row=r, column=1, value="Sin registros en el período.").font = Font(
            italic=True, color="94A3B8")
        r += 1

    if data.get("total"):
        ws.cell(row=r + 1, column=1, value=data["total"][0]).font = Font(bold=True, color="1E3A8A")
        tc = ws.cell(row=r + 1, column=ncols, value=data["total"][1])
        tc.font = Font(bold=True, color="1E3A8A")
        tc.alignment = Alignment(horizontal="right")

    for j in range(1, ncols + 1):
        width = max(
            [len(str(data["headers"][j - 1]))] +
            [len(str(row[j - 1])) for row in data["rows"] if j - 1 < len(row)] + [10]
        )
        ws.column_dimensions[ws.cell(row=hrow, column=j).column_letter].width = min(width + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
#  API pública
# --------------------------------------------------------------------------- #
def generar_archivo(reporte):
    """Devuelve (contenido_bytes, nombre_archivo, content_type)."""
    data = construir_datos(reporte)
    base = f"reporte_{reporte.tipo}_{reporte.id_reporte}"
    if reporte.formato == "Excel":
        contenido = _render_excel(data, reporte)
        return (
            contenido,
            f"{base}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    contenido = _render_pdf(data, reporte)
    return (contenido, f"{base}.pdf", "application/pdf")
